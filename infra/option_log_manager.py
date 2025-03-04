import os
from typing import Dict

import pandas as pd
from openpyxl.reader.excel import load_workbook

from model.enum.option_type import OptionType
from model.instrument.option import ETFOption
from model.instrument.option_series import OptionSeries


class OptionLogManager:
    def __init__(self, instrument_transform_full_symbol, base_dir=None):
        """初始化日志管理器"""
        self.base_dir = base_dir
        os.makedirs(self.base_dir, exist_ok=True)  # 创建根目录
        self.option_series : Dict[str, OptionSeries] = {}
        self.instrument_transform_full_symbol = instrument_transform_full_symbol

    def read_data_from_company(self, file_path: str) -> pd.DataFrame:
        """
        读取 国泰君安 数据并返回一个 DataFrame。

        :param file_path: 数据文件路径，CSV 格式
        :return: 数据的 DataFrame
        """

        print(f"read_data_from_company:{file_path}")

        # 读取数据文件，指定列名
        column_names = [
            'order_book_id', 'datetime',
            'a1', 'a2', 'a3', 'a4', 'a5',
            'b1', 'b2', 'b3', 'b4', 'b5',
            'a1_v', 'a2_v', 'a3_v', 'a4_v', 'a5_v',
            'b1_v', 'b2_v', 'b3_v', 'b4_v', 'b5_v'
        ]

        # 读取 CSV 文件到 pandas DataFrame，假设文件是 CSV 格式
        df = pd.read_csv(file_path, usecols=column_names)

        # 将 datetime 和 trading_date 转换为合适的格式
        df['datetime'] = pd.to_datetime(df['datetime'], errors='coerce')  # 将 datetime 列转换为日期时间

        # 过滤出时间段在 9:30-11:30 或 13:00-15:00 之间的数据
        df['time'] = df['datetime'].dt.time
        morning_start = pd.to_datetime("09:30:00").time()
        morning_end = pd.to_datetime("11:30:00").time()
        afternoon_start = pd.to_datetime("13:00:00").time()
        afternoon_end = pd.to_datetime("15:00:00").time()

        # 保留符合条件的时间段数据
        df = df[(df['time'] >= morning_start) & (df['time'] <= morning_end) |
                (df['time'] >= afternoon_start) & (df['time'] <= afternoon_end)]

        # 去掉时间列
        df.drop(columns='time', inplace=True)

        df = df.sort_values(by=['order_book_id', 'datetime'])

        # 设置 datetime 列为索引
        df.set_index('datetime', inplace=True)

        # 按 order_book_id 分组，然后对每个组进行 5 分钟重采样
        def resample_group(group):
            # 对该组进行重采样
            group_resampled = group.resample('5T').first()
            return group_resampled

        # 对每个 order_book_id 分组，分别进行重采样
        df_resampled = df.groupby('order_book_id').apply(resample_group)

        # 去掉缺失值的行
        df_resampled = df_resampled.dropna()

        return df_resampled

    def record_wing_para(self, option_series: OptionSeries):
        symbol = option_series.symbol
        print(f"record {symbol} option series wing para log.")
        # 目录和文件路径
        underlying_folder = os.path.join(self.base_dir, symbol[:-8])
        os.makedirs(underlying_folder, exist_ok=True)  # 确保目录存在
        file_path = os.path.join(underlying_folder, f"{option_series.expired_date}.xlsx")

        record_data = [
            [option_series.wing_model_para.time,
             option_series.wing_model_para.k1,
             option_series.wing_model_para.k2,
             option_series.wing_model_para.b,
             option_series.wing_model_para.v]
        ]
        df_new = pd.DataFrame(record_data, columns=[
            'timestamp',
            'k1',
            'k2',
            'b',
            'v',
        ])

        print(record_data)

        sheet_name = "General"
        if os.path.exists(file_path):
            wb = load_workbook(file_path)

            sheet_exists = sheet_name in wb.sheetnames  # 检查 Sheet 是否存在

            with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                if sheet_exists:
                    existing_df = pd.read_excel(file_path, sheet_name=sheet_name, engine="openpyxl")
                    df_combined = pd.concat([existing_df, df_new], ignore_index=True)
                else:
                    df_combined = df_new

                df_combined.to_excel(writer, sheet_name=sheet_name, index=False)
        else:
            # 文件不存在，创建新文件
            with pd.ExcelWriter(file_path, engine="openpyxl", mode="w") as writer:
                df_new.to_excel(writer, sheet_name=sheet_name, index=False)

        print(f"记录完成: {file_path} -> {sheet_name} Sheet")

    def record_option_log(self, option_series: OptionSeries):
        symbol = option_series.symbol
        print(f"record {symbol} option series log.")
        # 目录和文件路径
        underlying_folder = os.path.join(self.base_dir, symbol[:-8])
        os.makedirs(underlying_folder, exist_ok=True)  # 确保目录存在
        file_path = os.path.join(underlying_folder, f"{option_series.expired_date}.xlsx")

        # 用于存储所有的记录数据
        all_record_data = []

        # 遍历所有的 strike_price 和对应的 option_tuple
        for strike_price, option_tuple in option_series.strike_price_options.items():
            for option in option_tuple:
                if option is None:
                    continue
                # 添加每个 option 的数据
                record_data = [
                    [option.market_data.sending_time, option.strike_price, option.option_type,
                     option.market_data.bid_prices[0], option.market_data.ask_prices[0],
                     option.market_data.bid_volumes[0], option.market_data.ask_volumes[0]]
                ]
                all_record_data.extend(record_data)

        # 将所有的记录数据转为 DataFrame
        df_new = pd.DataFrame(all_record_data, columns=[
            'datetime', 'strike_price', 'option_type',
            'bid1_price', 'ask1_price', 'bid1_volume', 'ask1_volume'
        ])

        sheet_name = "OptionData"  # 所有数据放在一个sheet中

        # 如果文件存在
        if os.path.exists(file_path):
            # 打开文件并检查是否有该sheet
            wb = load_workbook(file_path)
            sheet_exists = sheet_name in wb.sheetnames

            with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                # 如果sheet存在，读取并追加数据
                if sheet_exists:
                    existing_df = pd.read_excel(file_path, sheet_name=sheet_name, engine="openpyxl")
                    df_combined = pd.concat([existing_df, df_new], ignore_index=True)  # 追加数据
                else:
                    df_combined = df_new

                # 将合并后的数据写入
                df_combined.to_excel(writer, sheet_name=sheet_name, index=False)
        else:
            # 文件不存在，创建新文件
            with pd.ExcelWriter(file_path, engine="openpyxl", mode="w") as writer:
                df_new.to_excel(writer, sheet_name=sheet_name, index=False)

        print(f"记录完成: {file_path} -> {sheet_name} Sheet")


if __name__ == '__main__':
    o = ETFOption("00000", "20250401", OptionType.C.name, 1000, "", "", 1)
    options = [o]
    o_series = OptionSeries("IF20250401", options)

    m = OptionLogManager()
    m.record_option_log(o_series)
